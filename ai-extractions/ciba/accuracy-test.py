#!/usr/bin/env python3
"""
JSON Ground Truth vs Prediction -> Validation & Accuracy Workbook
=================================================================

PURPOSE
-------
Compare a prediction JSON against a ground-truth JSON that follows this schema:

ROOT / Submission
├── root/account/broker/producer fields
└── propertyInfo[]                    # repeating LOCATION objects
    ├── location-level fields
    └── buildings[]                   # repeating BUILDING objects
        └── building-level fields

The script creates one .xlsx workbook with exactly two sheets:
    1. Validation
    2. Accuracy

CURRENT OUTPUT SCHEMA UPDATE
----------------------------
The current schema includes these additions relative to the previous version:
- Root / broker: brokerReferenceID
- Root / broker: brokerZipcode
- Location: excluded
- Location: verified

These fields are expected ground-truth values and are validated, color-coded,
reported, and included in accuracy calculations like all other scored fields.

KEY VALIDATION / ACCURACY RULES
-------------------------------
1. Ground truth is authoritative.
2. Prediction array order and source LocationNumber values do NOT define identity.
3. Locations are matched ONE-TO-ONE by actual content, primarily building addresses.
4. A predicted location cannot satisfy two separate ground-truth locations.
   - If prediction merges two expected locations, one may match and the other is
     considered missing as a separate location.
5. Buildings are matched only inside the matched parent location.
   - If the correct building was extracted but placed under the wrong parent
     location, its parent assignment is wrong.
6. The number of locations affects accuracy.
7. The number of buildings under each location affects accuracy.
8. Building-to-location membership affects accuracy.
9. Building field "LocationNumber" is informational only:
   - visible in Validation
   - NOT color-highlighted
   - NOT included in any accuracy calculation
10. Expected values/checks come from ground truth.
11. Extra prediction scalar fields are not scored unless they create a structural
    error (for example, extra locations/buildings through count mismatches).

CELL COLORS IN VALIDATION
-------------------------
Exact match:
    green  #b6d7a8

Same meaning, formatting-only difference:
    yellow #ffe599

Wrong / missing:
    red    #ea9999

DESIGN RULES
------------
- Font: Arial
- Font size: 10
- Horizontal alignment: left
- Vertical alignment: top
- Wrap text: on
- Section headers: #b7b7b7
- Table subheaders: #d9d9d9
- Accuracy Summary "Overall" row: #b6d7a8
- Sentence-case headers, noun/root word first where practical.
- No unexplained abbreviations such as GT or Pred.
- Use "ground truth" and "prediction".
- Report pointers such as L1, L2, B1 are only convenient references.
- "LocationNumber" from JSON is never used as identity.
- mansardRoof is written literally as null / true / false, never as a checkbox.

ACCURACY DEFINITIONS
--------------------
Accuracy (formatting accepted)
    = (Correct + Formatting issues) / Expected values or checks

Accuracy (strict)
    = Correct / Expected values or checks

Percentages are displayed to 4 decimal places.

STRUCTURAL CHECKS INCLUDED IN COMBINED ACCURACY
-----------------------------------------------
- overall location count
- each expected location exists as a separate predicted location
- building count under each expected location
- each expected building belongs to the correct matched parent location

The Accuracy sheet also reports "Overall data" and "Overall structure" separately
for transparency.

DEPENDENCY
----------
pip install openpyxl

EXAMPLE
-------
python validation_accuracy.py \
    --ground-truth "c10-ground-truth.json" \
    --prediction "c10-prediction.json" \
    --output "c10-validation.xlsx"
"""

from __future__ import annotations

import argparse
import itertools
import json
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# CONFIGURATION
# =============================================================================

SHEET_VALIDATION = "Validation"
SHEET_ACCURACY = "Accuracy"

COLOR_CORRECT = "B6D7A8"
COLOR_FORMATTING = "FFE599"
COLOR_WRONG = "EA9999"
COLOR_HEADER = "B7B7B7"
COLOR_SUBHEADER = "D9D9D9"
COLOR_BORDER = "D9D9D9"

FONT_NAME = "Arial"
FONT_SIZE = 10

# Informational field only. It remains visible, but it is never scored.
EXCLUDED_BUILDING_SCORE_FIELDS = {"LocationNumber"}


# Updated output schema (September 2026).
#
# Ground truth is normalized to this schema before validation so every expected
# field is always present, written to Validation, and included in accuracy.
# Prediction is reordered to the same schema but missing prediction fields are
# NOT filled with null; a missing prediction field must remain missing/wrong.
#
# Newly added fields in the current schema:
# - Root / broker: brokerReferenceID, brokerZipcode
# - Location: excluded, verified
ROOT_FIELD_ORDER = [
    "accountAddress1",
    "accountCity",
    "accountID",
    "accountName",
    "accountReferenceID",
    "accountState",
    "accountZipCode",
    "brokerCity",
    "brokerCompanyName",
    "brokerReferenceID",
    "brokerState",
    "brokerZipcode",
    "consumerCode",
    "effectiveDate",
    "expirationDate",
    "homeState",
    "producerFirstName",
    "producerLastName",
]

LOCATION_FIELD_ORDER = [
    "assistedLivingOperator",
    "buildingCount",
    "buildingValue",
    "city",
    "constructionTypeID",
    "countryID",
    "county",
    "direction",
    "excluded",
    "firstLineAddress",
    "housingTypeID",
    "industrialSqft",
    "latitude",
    "longitude",
    "name",
    "numberOfBuildings",
    "numberOfStories",
    "numberOfUnitsOrCondos",
    "occupancySubTypeID",
    "occupancyTypeID",
    "propertyCharacteristicsID",
    "propertyLocationID",
    "rentalValue",
    "residentialSqft",
    "retailSqft",
    "secondLineAddress",
    "sprinklerTypeID",
    "stateCode",
    "streetAddress1",
    "streetAddress2",
    "streetName",
    "streetNumber",
    "streetTypeID",
    "tenantScreeningTypeID",
    "tiv",
    "totalSqft",
    "unitNumber",
    "verified",
    "yearBuilt",
    "zipCode",
]

BUILDING_FIELD_ORDER = [
    "LocationNumber",
    "assistedLivingOperator",
    "buildingValue",
    "city",
    "constructionTypeID",
    "countryID",
    "county",
    "direction",
    "excluded",
    "firstLineAddress",
    "housingTypeID",
    "industrialSqft",
    "isPrimaryLimit",
    "latitude",
    "longitude",
    "mansardRoof",
    "name",
    "numBuilding",
    "numberOfBuildings",
    "numberOfStories",
    "numberOfUnitsOrCondos",
    "occupancySubTypeID",
    "occupancyTypeID",
    "propertyCharacteristicsID",
    "propertyLocationID",
    "rentalValue",
    "residentialSqft",
    "retailSqft",
    "roofEquipment",
    "secondLineAddress",
    "sprinklerTypeID",
    "stateCode",
    "streetAddress1",
    "streetAddress2",
    "streetName",
    "streetNumber",
    "streetTypeID",
    "tenantScreeningTypeID",
    "tiv",
    "totalSqft",
    "unitNumber",
    "verified",
    "yearBuilt",
    "zipCode",
]

NULL_SYNONYMS = {
    "",
    "null",
    "none",
    "nil",
    "n/a",
    "na",
    "n.a.",
    "not applicable",
    "not available",
    "unknown",
    "missing",
}

TRUE_SYNONYMS = {"true", "yes", "y"}
FALSE_SYNONYMS = {"false", "no", "n"}

ADDRESS_FIELDS = {
    "accountAddress1",
    "firstLineAddress",
    "secondLineAddress",
    "streetAddress1",
    "streetAddress2",
}

DATE_FIELDS = {
    "effectiveDate",
    "expirationDate",
}

# Location-level aggregates that often become wrong when a location is merged,
# split, or otherwise structurally mis-grouped.
AGGREGATE_LOCATION_FIELDS = {
    "buildingCount",
    "buildingValue",
    "numberOfBuildings",
    "numberOfUnitsOrCondos",
    "rentalValue",
    "tiv",
    "totalSqft",
    "industrialSqft",
    "residentialSqft",
    "retailSqft",
    "numberOfStories",
    "yearBuilt",
}


# =============================================================================
# DATA MODEL
# =============================================================================

MISSING = object()


@dataclass
class Record:
    record_type: str              # "Data" or "Structure"
    scope: str
    identity: str
    location_idx: Optional[int]
    building_idx: Optional[int]
    field: str
    reference: str
    ground_truth: Any
    prediction: Any
    status: str                   # Correct / Formatting Issue / Wrong


# =============================================================================
# OUTPUT SCHEMA NORMALIZATION
# =============================================================================

def _ordered_fields(
    source: Dict[str, Any],
    field_order: Sequence[str],
    *,
    fill_missing: bool,
) -> Dict[str, Any]:
    """
    Return a new dictionary in schema order.

    Ground truth uses fill_missing=True so every schema field is an expected
    value/check even if an older ground-truth file omitted that key.

    Prediction uses fill_missing=False so an omitted prediction field remains
    omitted and is scored as missing/wrong rather than being converted to null.
    """
    ordered: Dict[str, Any] = {}

    for field in field_order:
        if field in source:
            ordered[field] = source[field]
        elif fill_missing:
            ordered[field] = None

    # Preserve unexpected/additional fields after the known schema fields.
    for field, value in source.items():
        if field not in ordered and field not in {"propertyInfo", "buildings"}:
            ordered[field] = value

    return ordered


def normalize_output_schema(
    data: Dict[str, Any],
    *,
    fill_missing: bool,
) -> Dict[str, Any]:
    """
    Normalize a loaded JSON object to the current output schema.

    Current additions explicitly supported:
    - brokerReferenceID
    - brokerZipcode
    - propertyInfo[].excluded
    - propertyInfo[].verified

    Because downstream validation and accuracy logic iterates normalized
    ground-truth fields, these fields are automatically:
    - written in Validation,
    - compared against Prediction,
    - color-highlighted,
    - included in issue reporting,
    - included in per-location/root/overall accuracy calculations.
    """
    normalized = _ordered_fields(
        data,
        ROOT_FIELD_ORDER,
        fill_missing=fill_missing,
    )

    normalized_locations: List[Dict[str, Any]] = []

    for location in data.get("propertyInfo", []) or []:
        normalized_location = _ordered_fields(
            location,
            LOCATION_FIELD_ORDER,
            fill_missing=fill_missing,
        )

        normalized_buildings: List[Dict[str, Any]] = []
        for building in location.get("buildings", []) or []:
            normalized_building = _ordered_fields(
                building,
                BUILDING_FIELD_ORDER,
                fill_missing=fill_missing,
            )
            normalized_buildings.append(normalized_building)

        # Keep buildings[] in the established schema position:
        # assistedLivingOperator, buildingCount, buildingValue, buildings[], ...
        location_with_buildings: Dict[str, Any] = {}

        for field in ("assistedLivingOperator", "buildingCount", "buildingValue"):
            if field in normalized_location:
                location_with_buildings[field] = normalized_location[field]

        location_with_buildings["buildings"] = normalized_buildings

        for field in LOCATION_FIELD_ORDER:
            if field not in {"assistedLivingOperator", "buildingCount", "buildingValue"}:
                if field in normalized_location:
                    location_with_buildings[field] = normalized_location[field]

        # Preserve any additional location fields after the official schema.
        for field, value in normalized_location.items():
            if field not in location_with_buildings:
                location_with_buildings[field] = value

        normalized_locations.append(location_with_buildings)

    normalized["propertyInfo"] = normalized_locations
    return normalized


# =============================================================================
# BASIC HELPERS
# =============================================================================

def load_json(path: str | Path) -> Dict[str, Any]:
    path = Path(path)
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, dict):
        raise ValueError(f"{path.name}: root JSON value must be an object.")

    property_info = data.get("propertyInfo", [])
    if property_info is None:
        property_info = []
    if not isinstance(property_info, list):
        raise ValueError(f"{path.name}: propertyInfo must be an array/list.")

    for i, location in enumerate(property_info):
        if not isinstance(location, dict):
            raise ValueError(
                f"{path.name}: propertyInfo[{i}] must be an object."
            )
        buildings = location.get("buildings", [])
        if buildings is None:
            buildings = []
        if not isinstance(buildings, list):
            raise ValueError(
                f"{path.name}: propertyInfo[{i}].buildings must be an array/list."
            )
        for j, building in enumerate(buildings):
            if not isinstance(building, dict):
                raise ValueError(
                    f"{path.name}: propertyInfo[{i}].buildings[{j}] must be an object."
                )

    return data


def is_nullish(value: Any) -> bool:
    if value is None:
        return True
    return (
        isinstance(value, str)
        and value.strip().casefold() in NULL_SYNONYMS
    )


def normalize_text(value: Any) -> str:
    s = str(value).strip().casefold()
    s = s.replace("–", "-").replace("—", "-").replace("’", "'")
    s = re.sub(r"\s+", " ", s)
    # Ignore spaces immediately around ordinary punctuation.
    s = re.sub(r"\s*([.,;:#/\-])\s*", r"\1", s)
    return s


def normalize_address(value: Any) -> str:
    """
    Used for object matching and for address formatting-equivalence only.
    It is deliberately more permissive than normal text comparison.
    """
    if value is None:
        return ""

    s = str(value).strip().casefold()
    s = s.replace(".", "").replace(",", "")
    s = s.replace("#", " #")

    replacements = {
        r"\bavenue\b": "ave",
        r"\bstreet\b": "st",
        r"\broad\b": "rd",
        r"\bboulevard\b": "blvd",
        r"\bcourt\b": "ct",
        r"\bdrive\b": "dr",
        r"\blane\b": "ln",
        r"\bhighway\b": "hwy",
        r"\bparkway\b": "pkwy",
        r"\bplace\b": "pl",
        r"\bterrace\b": "ter",
        r"\bcircle\b": "cir",
    }
    for pattern, replacement in replacements.items():
        s = re.sub(pattern, replacement, s)

    s = re.sub(r"\s+", " ", s)
    return s.strip()


def try_parse_number(value: Any) -> Optional[float]:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    if not isinstance(value, str):
        return None

    s = value.strip()
    if not s:
        return None

    # Permit ordinary formatting-only characters.
    s = s.replace(",", "").replace("$", "")
    s = s.replace("%", "")
    try:
        return float(s)
    except ValueError:
        return None


def try_parse_date(value: Any) -> Optional[datetime]:
    if not isinstance(value, str):
        return None

    s = value.strip()
    if not s:
        return None

    formats = [
        "%m/%d/%Y",
        "%Y-%m-%d",
        "%m-%d-%Y",
        "%Y/%m/%d",
        "%m/%d/%y",
        "%Y.%m.%d",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def exact_same(a: Any, b: Any) -> bool:
    """
    Strict JSON-like comparison.
    bool is intentionally distinct from int.
    """
    if type(a) is not type(b):
        return False
    return a == b


def formatting_equivalent(field: str, a: Any, b: Any) -> bool:
    """
    Return True only when values communicate the same meaning and differ
    in representation/formatting.
    """
    if exact_same(a, b):
        return False

    # Null-like text.
    if is_nullish(a) and is_nullish(b):
        return True

    # Boolean text.
    if isinstance(a, bool) and isinstance(b, str):
        b_norm = b.strip().casefold()
        return (
            (a and b_norm in TRUE_SYNONYMS)
            or ((not a) and b_norm in FALSE_SYNONYMS)
        )
    if isinstance(b, bool) and isinstance(a, str):
        a_norm = a.strip().casefold()
        return (
            (b and a_norm in TRUE_SYNONYMS)
            or ((not b) and a_norm in FALSE_SYNONYMS)
        )

    # Number vs number-string or numeric formatting.
    a_num = try_parse_number(a)
    b_num = try_parse_number(b)
    if a_num is not None and b_num is not None:
        if math.isclose(a_num, b_num, rel_tol=0.0, abs_tol=1e-12):
            return True

    # Date formatting differences.
    if field in DATE_FIELDS:
        a_date = try_parse_date(a)
        b_date = try_parse_date(b)
        if a_date is not None and b_date is not None and a_date.date() == b_date.date():
            return True

    # Address abbreviations/punctuation/spacing.
    if field in ADDRESS_FIELDS and isinstance(a, str) and isinstance(b, str):
        return normalize_address(a) == normalize_address(b)

    # General case/spacing/punctuation differences.
    if isinstance(a, str) and isinstance(b, str):
        return normalize_text(a) == normalize_text(b)

    return False


def compare_status(field: str, ground_truth: Any, prediction: Any) -> str:
    if prediction is MISSING:
        return "Wrong"
    if exact_same(ground_truth, prediction):
        return "Correct"
    if formatting_equivalent(field, ground_truth, prediction):
        return "Formatting Issue"
    return "Wrong"


def display_value(field: str, value: Any) -> Any:
    if value is MISSING:
        return "<missing>"

    if value is None:
        return "null"

    # Explicit requirement: mansardRoof must be literal text, not a checkbox.
    if field == "mansardRoof":
        if value is True:
            return "true"
        if value is False:
            return "false"
        return str(value)

    # Keep booleans human-readable consistently.
    if value is True:
        return "true"
    if value is False:
        return "false"

    return value


def status_fill_hex(status: str) -> Optional[str]:
    if status == "Correct":
        return COLOR_CORRECT
    if status == "Formatting Issue":
        return COLOR_FORMATTING
    if status == "Wrong":
        return COLOR_WRONG
    return None


# =============================================================================
# LOCATION / BUILDING IDENTITY AND MATCHING
# =============================================================================

def building_address_set(location: Dict[str, Any]) -> set[str]:
    result: set[str] = set()
    for building in location.get("buildings", []) or []:
        address = normalize_address(building.get("streetAddress1"))
        if address:
            result.add(address)
    return result


def location_identity(location: Dict[str, Any]) -> str:
    """
    Friendly, human-readable identity used in Accuracy.
    Prefer the actual buildings under the location.
    """
    addresses = [
        str(b.get("streetAddress1")).strip()
        for b in (location.get("buildings", []) or [])
        if b.get("streetAddress1") not in (None, "")
    ]
    if addresses:
        return " | ".join(addresses)

    if location.get("streetAddress1"):
        return str(location["streetAddress1"])

    parts = [
        location.get("name"),
        location.get("city"),
        location.get("stateCode"),
        location.get("zipCode"),
    ]
    parts = [str(x) for x in parts if x not in (None, "")]
    return " | ".join(parts) if parts else "(location not identifiable by address)"


def building_identity(building: Dict[str, Any]) -> str:
    if building.get("streetAddress1"):
        return str(building["streetAddress1"])

    parts = [
        building.get("streetNumber"),
        building.get("streetName"),
        building.get("city"),
        building.get("stateCode"),
        building.get("zipCode"),
        building.get("name"),
    ]
    parts = [str(x) for x in parts if x not in (None, "")]
    return " | ".join(parts) if parts else "(building not identifiable by address)"


def location_pair_score(gt_loc: Dict[str, Any], pred_loc: Dict[str, Any]) -> float:
    """
    Location numbers are never used.

    Building-address overlap dominates because buildings define the actual
    composition of the location. Location-level fields only assist matching.
    """
    gt_addresses = building_address_set(gt_loc)
    pred_addresses = building_address_set(pred_loc)

    overlap = len(gt_addresses & pred_addresses)

    score = 0.0

    if overlap:
        coverage = overlap / max(1, len(gt_addresses))
        precision = overlap / max(1, len(pred_addresses))

        score += overlap * 100.0
        score += coverage * 30.0
        score += precision * 20.0
        score -= abs(len(gt_addresses) - len(pred_addresses)) * 2.0

    # Location main address.
    gt_main = normalize_address(gt_loc.get("streetAddress1"))
    pred_main = normalize_address(pred_loc.get("streetAddress1"))
    if gt_main and pred_main and gt_main == pred_main:
        score += 25.0

    # Supplemental location identity.
    supplemental = [
        ("city", 5.0),
        ("stateCode", 5.0),
        ("zipCode", 8.0),
        ("county", 3.0),
        ("name", 3.0),
    ]
    for field, weight in supplemental:
        gv = gt_loc.get(field)
        pv = pred_loc.get(field)
        if gv not in (None, "") and pv not in (None, ""):
            if normalize_text(gv) == normalize_text(pv):
                score += weight

    # No building overlap AND no meaningful address/location agreement means
    # it should not be treated as a candidate match.
    if overlap == 0 and score < 20.0:
        return -1000.0

    return score


def building_pair_score(gt_bld: Dict[str, Any], pred_bld: Dict[str, Any]) -> float:
    """
    Building matching happens only inside an already matched parent location.
    LocationNumber is ignored.
    """
    score = 0.0

    gt_addr = normalize_address(gt_bld.get("streetAddress1"))
    pred_addr = normalize_address(pred_bld.get("streetAddress1"))

    if gt_addr and pred_addr:
        if gt_addr == pred_addr:
            score += 200.0
        else:
            # Strongly discourage two different explicit addresses.
            score -= 100.0

    fields = [
        ("streetNumber", 25.0),
        ("streetName", 25.0),
        ("zipCode", 20.0),
        ("city", 10.0),
        ("stateCode", 8.0),
        ("yearBuilt", 10.0),
        ("name", 5.0),
        ("numberOfUnitsOrCondos", 5.0),
        ("totalSqft", 5.0),
    ]
    for field, weight in fields:
        gv = gt_bld.get(field)
        pv = pred_bld.get(field)

        if gv in (None, "") or pv in (None, ""):
            continue

        if exact_same(gv, pv) or formatting_equivalent(field, gv, pv):
            score += weight

    if score < 30.0:
        return -1000.0

    return score


def hungarian_maximize(score_matrix: Sequence[Sequence[float]]) -> List[Optional[int]]:
    """
    Pure-Python maximum-weight one-to-one assignment.

    Returns row -> column assignments.
    Dummy rows/columns are allowed by padding the matrix with zero-score cells.
    A caller should reject real-real matches below its semantic threshold.

    This avoids requiring scipy.
    """
    rows = len(score_matrix)
    cols = len(score_matrix[0]) if rows else 0

    if rows == 0:
        return []
    if cols == 0:
        return [None] * rows

    n = max(rows, cols)

    padded = [[0.0] * n for _ in range(n)]
    for i in range(rows):
        for j in range(cols):
            padded[i][j] = float(score_matrix[i][j])

    max_score = max(0.0, max(max(row) for row in padded))
    cost = [[max_score - value for value in row] for row in padded]

    # Hungarian algorithm for minimum cost, 1-indexed.
    u = [0.0] * (n + 1)
    v = [0.0] * (n + 1)
    p = [0] * (n + 1)
    way = [0] * (n + 1)

    for i in range(1, n + 1):
        p[0] = i
        j0 = 0
        minv = [float("inf")] * (n + 1)
        used = [False] * (n + 1)

        while True:
            used[j0] = True
            i0 = p[j0]
            delta = float("inf")
            j1 = 0

            for j in range(1, n + 1):
                if used[j]:
                    continue

                cur = cost[i0 - 1][j - 1] - u[i0] - v[j]
                if cur < minv[j]:
                    minv[j] = cur
                    way[j] = j0

                if minv[j] < delta:
                    delta = minv[j]
                    j1 = j

            for j in range(0, n + 1):
                if used[j]:
                    u[p[j]] += delta
                    v[j] -= delta
                else:
                    minv[j] -= delta

            j0 = j1
            if p[j0] == 0:
                break

        while True:
            j1 = way[j0]
            p[j0] = p[j1]
            j0 = j1
            if j0 == 0:
                break

    row_to_col: List[Optional[int]] = [None] * n
    for j in range(1, n + 1):
        if p[j] != 0:
            row_to_col[p[j] - 1] = j - 1

    return row_to_col[:rows]


def match_locations(
    gt_locations: Sequence[Dict[str, Any]],
    pred_locations: Sequence[Dict[str, Any]],
) -> Tuple[Dict[int, Optional[int]], Dict[int, int]]:
    if not gt_locations:
        return {}, {}

    if not pred_locations:
        return {i: None for i in range(len(gt_locations))}, {}

    matrix = [
        [location_pair_score(g, p) for p in pred_locations]
        for g in gt_locations
    ]

    assignment = hungarian_maximize(matrix)

    gt_to_pred: Dict[int, Optional[int]] = {}
    pred_to_gt: Dict[int, int] = {}

    for gi, pi in enumerate(assignment):
        if (
            pi is not None
            and pi < len(pred_locations)
            and matrix[gi][pi] > 0
        ):
            gt_to_pred[gi] = pi
            pred_to_gt[pi] = gi
        else:
            gt_to_pred[gi] = None

    return gt_to_pred, pred_to_gt


def match_buildings_within_location(
    gt_buildings: Sequence[Dict[str, Any]],
    pred_buildings: Sequence[Dict[str, Any]],
) -> Dict[int, Optional[int]]:
    if not gt_buildings:
        return {}

    if not pred_buildings:
        return {i: None for i in range(len(gt_buildings))}

    matrix = [
        [building_pair_score(g, p) for p in pred_buildings]
        for g in gt_buildings
    ]
    assignment = hungarian_maximize(matrix)

    result: Dict[int, Optional[int]] = {}
    for gi, pi in enumerate(assignment):
        if (
            pi is not None
            and pi < len(pred_buildings)
            and matrix[gi][pi] > 0
        ):
            result[gi] = pi
        else:
            result[gi] = None

    return result


# =============================================================================
# SCORING RECORDS
# =============================================================================

def build_records(
    gt: Dict[str, Any],
    pred: Dict[str, Any],
) -> Tuple[
    List[Record],
    List[Record],
    Dict[int, Optional[int]],
    Dict[int, int],
    Dict[Tuple[int, int], Optional[int]],
    Dict[str, List[Tuple[int, int]]],
]:
    gt_locations = gt.get("propertyInfo", []) or []
    pred_locations = pred.get("propertyInfo", []) or []

    gt_to_pred, pred_to_gt = match_locations(gt_locations, pred_locations)

    building_matches: Dict[Tuple[int, int], Optional[int]] = {}

    # Index every predicted building globally by normalized address. This is
    # used only to explain wrong-parent placement, never to grant data credit.
    global_pred_buildings: Dict[str, List[Tuple[int, int]]] = defaultdict(list)
    for pi, ploc in enumerate(pred_locations):
        for pbi, pb in enumerate(ploc.get("buildings", []) or []):
            address = normalize_address(pb.get("streetAddress1"))
            if address:
                global_pred_buildings[address].append((pi, pbi))

    for gi, gloc in enumerate(gt_locations):
        pi = gt_to_pred.get(gi)
        gt_buildings = gloc.get("buildings", []) or []
        pred_buildings = (
            pred_locations[pi].get("buildings", []) or []
            if pi is not None
            else []
        )

        local_matches = match_buildings_within_location(
            gt_buildings,
            pred_buildings,
        )
        for bi, pbi in local_matches.items():
            building_matches[(gi, bi)] = pbi

    data_records: List[Record] = []
    structural_records: List[Record] = []

    # Root data.
    for field, ground_truth_value in gt.items():
        if field == "propertyInfo":
            continue

        prediction_value = pred.get(field, MISSING)

        data_records.append(
            Record(
                record_type="Data",
                scope="Root",
                identity="Submission / Root",
                location_idx=None,
                building_idx=None,
                field=field,
                reference=field,
                ground_truth=ground_truth_value,
                prediction=prediction_value,
                status=compare_status(
                    field,
                    ground_truth_value,
                    prediction_value,
                ),
            )
        )

    # Global structural check: location count.
    structural_records.append(
        Record(
            record_type="Structure",
            scope="Overall structure",
            identity="Locations",
            location_idx=None,
            building_idx=None,
            field="Location count",
            reference="Location count",
            ground_truth=len(gt_locations),
            prediction=len(pred_locations),
            status=(
                "Correct"
                if len(gt_locations) == len(pred_locations)
                else "Wrong"
            ),
        )
    )

    # Location + building data / structure.
    for gi, gloc in enumerate(gt_locations):
        pi = gt_to_pred.get(gi)
        ploc = pred_locations[pi] if pi is not None else MISSING

        gt_buildings = gloc.get("buildings", []) or []
        pred_buildings = (
            pred_locations[pi].get("buildings", []) or []
            if pi is not None
            else []
        )

        # Separate-location existence check.
        structural_records.append(
            Record(
                record_type="Structure",
                scope=f"L{gi + 1}",
                identity=location_identity(gloc),
                location_idx=gi,
                building_idx=None,
                field="Location exists as separate location",
                reference=f"Location: {location_identity(gloc)}",
                ground_truth="Present",
                prediction="Present" if pi is not None else "Missing",
                status="Correct" if pi is not None else "Wrong",
            )
        )

        # Building count under this location.
        structural_records.append(
            Record(
                record_type="Structure",
                scope=f"L{gi + 1}",
                identity=location_identity(gloc),
                location_idx=gi,
                building_idx=None,
                field="Buildings under location",
                reference=f"Buildings under L{gi + 1}",
                ground_truth=len(gt_buildings),
                prediction=len(pred_buildings),
                status=(
                    "Correct"
                    if len(gt_buildings) == len(pred_buildings)
                    else "Wrong"
                ),
            )
        )

        # Location-level data fields.
        for field, ground_truth_value in gloc.items():
            if field == "buildings":
                continue

            prediction_value = (
                MISSING
                if ploc is MISSING
                else ploc.get(field, MISSING)
            )

            data_records.append(
                Record(
                    record_type="Data",
                    scope=f"L{gi + 1}",
                    identity=location_identity(gloc),
                    location_idx=gi,
                    building_idx=None,
                    field=field,
                    reference=f"propertyInfo[{gi}].{field}",
                    ground_truth=ground_truth_value,
                    prediction=prediction_value,
                    status=compare_status(
                        field,
                        ground_truth_value,
                        prediction_value,
                    ),
                )
            )

        # Buildings.
        for bi, gb in enumerate(gt_buildings):
            pbi = building_matches.get((gi, bi))

            if (
                pi is not None
                and pbi is not None
                and pbi < len(pred_buildings)
            ):
                pb = pred_buildings[pbi]
            else:
                pb = MISSING

            gt_building_name = building_identity(gb)

            # Parent-membership structural check.
            if pb is not MISSING:
                membership_prediction = "Correct parent location"
                membership_status = "Correct"
            else:
                address = normalize_address(gb.get("streetAddress1"))
                elsewhere = global_pred_buildings.get(address, [])

                if elsewhere:
                    parent_pi, _ = elsewhere[0]
                    matched_gt = pred_to_gt.get(parent_pi)

                    if matched_gt is None:
                        membership_prediction = f"Found under P{parent_pi + 1}"
                    else:
                        membership_prediction = (
                            f"Found under P{parent_pi + 1} / "
                            f"matched L{matched_gt + 1}"
                        )
                else:
                    membership_prediction = "Missing"

                membership_status = "Wrong"

            structural_records.append(
                Record(
                    record_type="Structure",
                    scope=f"L{gi + 1} / B{bi + 1}",
                    identity=gt_building_name,
                    location_idx=gi,
                    building_idx=bi,
                    field="Building belongs to location",
                    reference=f"Parent location for {gt_building_name}",
                    ground_truth=f"Under L{gi + 1}",
                    prediction=membership_prediction,
                    status=membership_status,
                )
            )

            # Building data. LocationNumber is visible later in Validation,
            # but intentionally excluded here.
            for field, ground_truth_value in gb.items():
                if field in EXCLUDED_BUILDING_SCORE_FIELDS:
                    continue

                prediction_value = (
                    MISSING
                    if pb is MISSING
                    else pb.get(field, MISSING)
                )

                data_records.append(
                    Record(
                        record_type="Data",
                        scope=f"L{gi + 1} / B{bi + 1}",
                        identity=gt_building_name,
                        location_idx=gi,
                        building_idx=bi,
                        field=field,
                        reference=(
                            f"propertyInfo[{gi}]."
                            f"buildings[{bi}].{field}"
                        ),
                        ground_truth=ground_truth_value,
                        prediction=prediction_value,
                        status=compare_status(
                            field,
                            ground_truth_value,
                            prediction_value,
                        ),
                    )
                )

    return (
        data_records,
        structural_records,
        gt_to_pred,
        pred_to_gt,
        building_matches,
        global_pred_buildings,
    )


# =============================================================================
# EXPLANATORY REMARKS
# =============================================================================

def issue_remark(record: Record) -> str:
    if record.record_type == "Structure":
        if record.field == "Location count":
            return "Expected and predicted location counts differ."

        if record.field == "Location exists as separate location":
            if record.status == "Wrong":
                return (
                    "Ground truth location was merged or not separately "
                    "predicted."
                )

        if record.field == "Buildings under location":
            return "Wrong number of buildings assigned to this location."

        if record.field == "Building belongs to location":
            return (
                "Building found under wrong parent location or missing."
            )

        return ""

    if record.status == "Formatting Issue":
        return "Formatting only; same meaning."

    if record.prediction is MISSING:
        return (
            "Value missing because matched location/building is absent."
        )

    if (
        record.location_idx is not None
        and record.building_idx is None
        and record.field in AGGREGATE_LOCATION_FIELDS
    ):
        return (
            "Location-level value differs; grouping may affect aggregate."
        )

    if record.field == "numBuilding":
        return "Building sequence number differs."

    if record.ground_truth is None and record.prediction is not None:
        return "Expected null; prediction populated a value."

    if record.ground_truth is not None and record.prediction is None:
        return "Prediction returned null."

    if (
        isinstance(record.ground_truth, (int, float))
        and not isinstance(record.ground_truth, bool)
        and isinstance(record.prediction, (int, float))
        and not isinstance(record.prediction, bool)
    ):
        return "Numeric value differs."

    if (
        isinstance(record.ground_truth, str)
        and isinstance(record.prediction, str)
    ):
        return "Text value differs."

    return ""


# =============================================================================
# WORKBOOK STYLING
# =============================================================================

THIN_SIDE = Side(style="thin", color=COLOR_BORDER)
TABLE_BORDER = Border(bottom=THIN_SIDE, right=THIN_SIDE)


def apply_common_cell_style(cell, bold: bool = False) -> None:
    cell.font = Font(
        name=FONT_NAME,
        size=FONT_SIZE,
        bold=bold,
    )
    cell.alignment = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True,
    )


def fill_cell(cell, hex_color: Optional[str]) -> None:
    if not hex_color:
        return
    cell.fill = PatternFill(
        fill_type="solid",
        fgColor=hex_color,
    )


def style_section_header(
    ws,
    row: int,
    start_col: int,
    end_col: int,
) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        apply_common_cell_style(cell, bold=True)
        fill_cell(cell, COLOR_HEADER)


def style_subheader(
    ws,
    row: int,
    start_col: int,
    end_col: int,
) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        apply_common_cell_style(cell, bold=True)
        fill_cell(cell, COLOR_SUBHEADER)
        cell.border = TABLE_BORDER


def apply_global_styles(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            apply_common_cell_style(
                cell,
                bold=bool(cell.font.bold),
            )


def apply_table_borders(
    ws,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
) -> None:
    if end_row < start_row:
        return

    for row in ws.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col,
    ):
        for cell in row:
            cell.border = TABLE_BORDER


def set_column_widths(
    ws,
    widths: Dict[int, float],
) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# =============================================================================
# VALIDATION SHEET
# =============================================================================

def write_validation_sheet(
    ws,
    gt: Dict[str, Any],
    pred: Dict[str, Any],
    gt_to_pred: Dict[int, Optional[int]],
    building_matches: Dict[Tuple[int, int], Optional[int]],
) -> None:
    gt_locations = gt.get("propertyInfo", []) or []
    pred_locations = pred.get("propertyInfo", []) or []

    # Root / submission section.
    ws["A1"] = "Root / submission"

    root_fields = [
        (field, value)
        for field, value in gt.items()
        if field != "propertyInfo"
    ]

    # Determine maximum required width dynamically from the largest building set.
    max_buildings = max(
        [len(loc.get("buildings", []) or []) for loc in gt_locations]
        + [0]
    )
    max_columns = max(2, 2 + max_buildings * 2)

    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=max_columns,
    )
    style_section_header(ws, 1, 1, max_columns)

    ws["A2"] = "Root field"
    ws["B2"] = "Value"
    style_subheader(ws, 2, 1, 2)

    row = 3
    for field, ground_truth_value in root_fields:
        prediction_value = pred.get(field, MISSING)
        status = compare_status(
            field,
            ground_truth_value,
            prediction_value,
        )

        ws.cell(row=row, column=1, value=field)
        ws.cell(
            row=row,
            column=2,
            value=display_value(field, ground_truth_value),
        )

        apply_common_cell_style(ws.cell(row=row, column=1), bold=True)
        apply_common_cell_style(ws.cell(row=row, column=2))

        fill_cell(
            ws.cell(row=row, column=2),
            status_fill_hex(status),
        )
        row += 1

    # propertyInfo section.
    ws.cell(
        row=row,
        column=1,
        value=(
            f"propertyInfo[] — Ground truth: {len(gt_locations)} locations "
            f"| Prediction: {len(pred_locations)} locations"
        ),
    )
    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=max_columns,
    )
    style_section_header(ws, row, 1, max_columns)
    row += 2

    # Each GT location.
    for gi, gloc in enumerate(gt_locations):
        pi = gt_to_pred.get(gi)
        ploc = pred_locations[pi] if pi is not None else MISSING

        gt_identity = location_identity(gloc)
        if pi is None:
            pred_identity = "No separate predicted location"
        else:
            pred_identity = (
                f"P{pi + 1}: {location_identity(pred_locations[pi])}"
            )

        ws.cell(
            row=row,
            column=1,
            value=(
                f"L{gi + 1} (pointer only) — "
                f"Location (ground truth): {gt_identity} — "
                f"Location (prediction): {pred_identity}"
            ),
        )
        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=max_columns,
        )
        style_section_header(ws, row, 1, max_columns)
        row += 1

        # Subheaders.
        ws.cell(row=row, column=1, value="Location field")
        ws.cell(row=row, column=2, value="Value")
        style_subheader(ws, row, 1, 2)

        gt_buildings = gloc.get("buildings", []) or []

        for bi in range(len(gt_buildings)):
            field_col = 3 + bi * 2
            value_col = field_col + 1

            ws.cell(
                row=row,
                column=field_col,
                value=f"Building field (B{bi + 1})",
            )
            ws.cell(
                row=row,
                column=value_col,
                value="Value",
            )
            style_subheader(
                ws,
                row,
                field_col,
                value_col,
            )

        row += 1
        data_start_row = row

        location_fields = [
            (field, value)
            for field, value in gloc.items()
            if field != "buildings"
        ]

        max_field_rows = max(
            [len(location_fields)]
            + [
                len(building)
                for building in gt_buildings
            ]
            + [0]
        )

        # Location fields.
        for offset, (field, ground_truth_value) in enumerate(location_fields):
            target_row = data_start_row + offset
            prediction_value = (
                MISSING
                if ploc is MISSING
                else ploc.get(field, MISSING)
            )
            status = compare_status(
                field,
                ground_truth_value,
                prediction_value,
            )

            ws.cell(
                row=target_row,
                column=1,
                value=field,
            )
            ws.cell(
                row=target_row,
                column=2,
                value=display_value(
                    field,
                    ground_truth_value,
                ),
            )

            apply_common_cell_style(
                ws.cell(row=target_row, column=1),
                bold=True,
            )
            apply_common_cell_style(
                ws.cell(row=target_row, column=2),
            )
            fill_cell(
                ws.cell(row=target_row, column=2),
                status_fill_hex(status),
            )

        # Building fields.
        for bi, gb in enumerate(gt_buildings):
            pbi = building_matches.get((gi, bi))

            if (
                pi is not None
                and pbi is not None
                and pbi < len(
                    pred_locations[pi].get("buildings", []) or []
                )
            ):
                pb = (
                    pred_locations[pi]
                    .get("buildings", [])[pbi]
                )
            else:
                pb = MISSING

            field_col = 3 + bi * 2
            value_col = field_col + 1

            for offset, (field, ground_truth_value) in enumerate(gb.items()):
                target_row = data_start_row + offset

                ws.cell(
                    row=target_row,
                    column=field_col,
                    value=field,
                )
                ws.cell(
                    row=target_row,
                    column=value_col,
                    value=display_value(
                        field,
                        ground_truth_value,
                    ),
                )

                apply_common_cell_style(
                    ws.cell(
                        row=target_row,
                        column=field_col,
                    ),
                    bold=True,
                )
                apply_common_cell_style(
                    ws.cell(
                        row=target_row,
                        column=value_col,
                    )
                )

                # Explicit rule: LocationNumber is not scored/highlighted.
                if field not in EXCLUDED_BUILDING_SCORE_FIELDS:
                    prediction_value = (
                        MISSING
                        if pb is MISSING
                        else pb.get(field, MISSING)
                    )
                    status = compare_status(
                        field,
                        ground_truth_value,
                        prediction_value,
                    )
                    fill_cell(
                        ws.cell(
                            row=target_row,
                            column=value_col,
                        ),
                        status_fill_hex(status),
                    )

        # Light bottom border across this location block.
        block_end = data_start_row + max_field_rows - 1
        apply_table_borders(
            ws,
            data_start_row,
            block_end,
            1,
            max_columns,
        )

        row = block_end + 2

    # Widths: field/value pairs repeat dynamically.
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 24

    for bi in range(max_buildings):
        field_col = 3 + bi * 2
        value_col = field_col + 1
        ws.column_dimensions[get_column_letter(field_col)].width = 28
        ws.column_dimensions[get_column_letter(value_col)].width = 22

    apply_global_styles(ws)
    ws.freeze_panes = "A3"


# =============================================================================
# ACCURACY SHEET
# =============================================================================

def summarize_records(
    records: Sequence[Record],
    scope: str,
    identity: str,
    level: str,
    remark: str = "",
) -> List[Any]:
    counts = Counter(record.status for record in records)
    expected = len(records)
    correct = counts["Correct"]
    formatting = counts["Formatting Issue"]
    wrong = counts["Wrong"]

    accuracy_formatting = (
        (correct + formatting) / expected
        if expected
        else 0.0
    )
    accuracy_strict = (
        correct / expected
        if expected
        else 0.0
    )

    return [
        scope,
        identity,
        level,
        expected,
        correct,
        formatting,
        wrong,
        accuracy_formatting,
        accuracy_strict,
        remark,
    ]


def write_accuracy_sheet(
    ws,
    gt: Dict[str, Any],
    pred: Dict[str, Any],
    data_records: Sequence[Record],
    structural_records: Sequence[Record],
    gt_to_pred: Dict[int, Optional[int]],
    pred_to_gt: Dict[int, int],
    building_matches: Dict[Tuple[int, int], Optional[int]],
    global_pred_buildings: Dict[str, List[Tuple[int, int]]],
) -> None:
    gt_locations = gt.get("propertyInfo", []) or []
    pred_locations = pred.get("propertyInfo", []) or []
    all_records = list(data_records) + list(structural_records)

    row = 1

    # -------------------------------------------------------------------------
    # Structure
    # -------------------------------------------------------------------------
    ws.cell(row=row, column=1, value="Structure")
    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=9,
    )
    style_section_header(ws, row, 1, 9)
    row += 1

    structure_headers = [
        "Location (pointer)",
        "Location (ground truth)",
        "Location (prediction)",
        "Location status",
        "Buildings (expected)",
        "Buildings (prediction)",
        "Buildings (correct parent)",
        "Buildings (wrong parent / missing)",
        "Remarks",
    ]
    for col, value in enumerate(structure_headers, start=1):
        ws.cell(row=row, column=col, value=value)
    style_subheader(ws, row, 1, 9)
    row += 1

    structure_start = row

    for gi, gloc in enumerate(gt_locations):
        pi = gt_to_pred.get(gi)
        expected_buildings = len(gloc.get("buildings", []) or [])

        if pi is None:
            prediction_identity = "<missing>"
            predicted_buildings = 0
            correct_parent = 0
            wrong_parent_or_missing = expected_buildings
            location_status = "Wrong"
            remark = (
                "No separate predicted location; location was merged or missed."
            )
        else:
            ploc = pred_locations[pi]
            prediction_identity = location_identity(ploc)
            predicted_buildings = len(ploc.get("buildings", []) or [])

            correct_parent = sum(
                1
                for bi in range(expected_buildings)
                if building_matches.get((gi, bi)) is not None
            )
            wrong_parent_or_missing = (
                expected_buildings - correct_parent
            )

            location_status = (
                "Correct"
                if (
                    expected_buildings == predicted_buildings
                    and wrong_parent_or_missing == 0
                )
                else "Wrong"
            )

            if predicted_buildings > expected_buildings:
                remark = (
                    f"{predicted_buildings - expected_buildings} "
                    f"extra building(s) assigned to this predicted location."
                )
            elif predicted_buildings < expected_buildings:
                remark = (
                    f"{expected_buildings - predicted_buildings} "
                    f"building(s) missing from this predicted location."
                )
            else:
                remark = "Location and building membership match."

        values = [
            f"L{gi + 1}",
            location_identity(gloc),
            prediction_identity,
            location_status,
            expected_buildings,
            predicted_buildings,
            correct_parent,
            wrong_parent_or_missing,
            remark,
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    structure_end = row - 1
    apply_table_borders(
        ws,
        structure_start - 1,
        structure_end,
        1,
        9,
    )
    row += 1

    # -------------------------------------------------------------------------
    # Location and building count
    # -------------------------------------------------------------------------
    ws.cell(
        row=row,
        column=1,
        value="Location and building count",
    )
    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=5,
    )
    style_section_header(ws, row, 1, 5)
    row += 1

    count_headers = [
        "Locations (expected)",
        "Locations (prediction)",
        "Buildings (expected)",
        "Buildings (prediction)",
        "Remarks",
    ]
    for col, value in enumerate(count_headers, start=1):
        ws.cell(row=row, column=col, value=value)
    style_subheader(ws, row, 1, 5)
    row += 1

    expected_buildings_total = sum(
        len(loc.get("buildings", []) or [])
        for loc in gt_locations
    )
    predicted_buildings_total = sum(
        len(loc.get("buildings", []) or [])
        for loc in pred_locations
    )

    count_remark = ""
    if (
        len(gt_locations) != len(pred_locations)
        or expected_buildings_total != predicted_buildings_total
    ):
        count_remark = (
            "Location/building counts differ."
        )
    elif any(
        len((loc.get("buildings", []) or []))
        != len(
            (
                pred_locations[gt_to_pred[i]]
                .get("buildings", []) or []
            )
            if gt_to_pred.get(i) is not None
            else []
        )
        for i, loc in enumerate(gt_locations)
    ):
        count_remark = (
            "Overall building count matches, but locations/building "
            "assignments do not."
        )
    else:
        count_remark = "Counts match."

    count_values = [
        len(gt_locations),
        len(pred_locations),
        expected_buildings_total,
        predicted_buildings_total,
        count_remark,
    ]
    for col, value in enumerate(count_values, start=1):
        ws.cell(row=row, column=col, value=value)

    apply_table_borders(ws, row - 1, row, 1, 5)
    row += 2

    # -------------------------------------------------------------------------
    # Location overview
    # -------------------------------------------------------------------------
    ws.cell(row=row, column=1, value="Location overview")
    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=8,
    )
    style_section_header(ws, row, 1, 8)
    row += 1

    location_headers = [
        "Location (pointer)",
        "Location (ground truth)",
        "Location (prediction)",
        "Address (ground truth)",
        "Address (prediction)",
        "Location match",
        "Building assignment",
        "Remarks",
    ]
    for col, value in enumerate(location_headers, start=1):
        ws.cell(row=row, column=col, value=value)
    style_subheader(ws, row, 1, 8)
    row += 1

    location_overview_start = row

    for gi, gloc in enumerate(gt_locations):
        pi = gt_to_pred.get(gi)
        expected_count = len(gloc.get("buildings", []) or [])

        if pi is None:
            values = [
                f"L{gi + 1}",
                location_identity(gloc),
                "<missing>",
                display_value(
                    "streetAddress1",
                    gloc.get("streetAddress1"),
                ),
                "<missing>",
                "Wrong",
                "Wrong",
                "No separate predicted location.",
            ]
        else:
            ploc = pred_locations[pi]
            matched_buildings = sum(
                1
                for bi in range(expected_count)
                if building_matches.get((gi, bi)) is not None
            )

            location_match = "Correct"
            building_assignment = (
                "Correct"
                if (
                    matched_buildings == expected_count
                    and len(ploc.get("buildings", []) or [])
                    == expected_count
                )
                else "Wrong"
            )

            if building_assignment == "Correct":
                remark = (
                    "Location matched by actual building addresses; "
                    "source location numbers ignored."
                )
            else:
                remark = (
                    "Location matched, but wrong building count/membership."
                )

            values = [
                f"L{gi + 1}",
                location_identity(gloc),
                location_identity(ploc),
                display_value(
                    "streetAddress1",
                    gloc.get("streetAddress1"),
                ),
                display_value(
                    "streetAddress1",
                    ploc.get("streetAddress1"),
                ),
                location_match,
                building_assignment,
                remark,
            ]

        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    location_overview_end = row - 1
    apply_table_borders(
        ws,
        location_overview_start - 1,
        location_overview_end,
        1,
        8,
    )
    row += 1

    # -------------------------------------------------------------------------
    # Building overview
    # -------------------------------------------------------------------------
    ws.cell(row=row, column=1, value="Building overview")
    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=10,
    )
    style_section_header(ws, row, 1, 10)
    row += 1

    building_headers = [
        "Location (pointer)",
        "Location (ground truth)",
        "Building (pointer)",
        "Building (ground truth)",
        "Parent location (ground truth)",
        "Building (prediction)",
        "Parent location (prediction)",
        "Building found?",
        "Parent location correct?",
        "Remarks",
    ]
    for col, value in enumerate(building_headers, start=1):
        ws.cell(row=row, column=col, value=value)
    style_subheader(ws, row, 1, 10)
    row += 1

    building_overview_start = row

    for gi, gloc in enumerate(gt_locations):
        pi = gt_to_pred.get(gi)

        for bi, gb in enumerate(gloc.get("buildings", []) or []):
            gt_building = building_identity(gb)
            pbi = building_matches.get((gi, bi))

            if (
                pi is not None
                and pbi is not None
                and pbi < len(
                    pred_locations[pi].get("buildings", []) or []
                )
            ):
                pb = pred_locations[pi].get("buildings", [])[pbi]
                prediction_building = building_identity(pb)
                prediction_parent = location_identity(pred_locations[pi])
                found = "Yes"
                correct_parent = "Yes"
                remark = (
                    "Correct building under correct matched location."
                )
            else:
                address_key = normalize_address(
                    gb.get("streetAddress1")
                )
                elsewhere = global_pred_buildings.get(
                    address_key,
                    [],
                )

                if elsewhere:
                    parent_pi, other_pbi = elsewhere[0]
                    pb = (
                        pred_locations[parent_pi]
                        .get("buildings", [])[other_pbi]
                    )
                    prediction_building = building_identity(pb)
                    prediction_parent = location_identity(
                        pred_locations[parent_pi]
                    )
                    found = "Yes"
                    correct_parent = "No"
                    remark = (
                        "Building was extracted but assigned to the wrong "
                        "location."
                    )
                else:
                    prediction_building = "<missing>"
                    prediction_parent = "<missing>"
                    found = "No"
                    correct_parent = "No"
                    remark = "Building missing from prediction."

            values = [
                f"L{gi + 1}",
                location_identity(gloc),
                f"B{bi + 1}",
                gt_building,
                location_identity(gloc),
                prediction_building,
                prediction_parent,
                found,
                correct_parent,
                remark,
            ]
            for col, value in enumerate(values, start=1):
                ws.cell(row=row, column=col, value=value)
            row += 1

    building_overview_end = row - 1
    apply_table_borders(
        ws,
        building_overview_start - 1,
        building_overview_end,
        1,
        10,
    )
    row += 1

    # -------------------------------------------------------------------------
    # Accuracy summary
    # -------------------------------------------------------------------------
    ws.cell(row=row, column=1, value="Accuracy summary")
    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=10,
    )
    style_section_header(ws, row, 1, 10)
    row += 1

    summary_headers = [
        "Scope",
        "Location / building",
        "Level",
        "Values / checks (expected)",
        "Correct",
        "Formatting issues",
        "Wrong",
        "Accuracy (formatting accepted)",
        "Accuracy (strict)",
        "Remarks",
    ]
    for col, value in enumerate(summary_headers, start=1):
        ws.cell(row=row, column=col, value=value)
    style_subheader(ws, row, 1, 10)
    row += 1

    accuracy_summary_start = row
    summary_rows: List[List[Any]] = []

    # Overall combined.
    summary_rows.append(
        summarize_records(
            all_records,
            "Overall",
            "All scored data + structure",
            "Combined",
            (
                "Includes location count, location existence, building count, "
                "and building membership. LocationNumber excluded."
            ),
        )
    )

    # Transparent components.
    summary_rows.append(
        summarize_records(
            data_records,
            "Overall data",
            "All JSON values",
            "Data",
            "LocationNumber excluded.",
        )
    )
    summary_rows.append(
        summarize_records(
            structural_records,
            "Overall structure",
            "Locations / buildings",
            "Structure",
            (
                "Location count and parent-child membership are scored."
            ),
        )
    )

    # Root.
    root_records = [
        r
        for r in all_records
        if r.scope == "Root"
    ]
    summary_rows.append(
        summarize_records(
            root_records,
            "Root",
            "Submission / Root",
            "Data",
            "",
        )
    )

    # Locations and buildings.
    for gi, gloc in enumerate(gt_locations):
        loc_identity = location_identity(gloc)

        location_total = [
            r
            for r in all_records
            if r.location_idx == gi
        ]
        summary_rows.append(
            summarize_records(
                location_total,
                f"L{gi + 1}",
                loc_identity,
                "Location total",
                (
                    "Includes location structure, location fields, building "
                    "membership, and building data."
                ),
            )
        )

        location_fields_and_structure = [
            r
            for r in all_records
            if (
                r.location_idx == gi
                and r.building_idx is None
            )
        ]
        summary_rows.append(
            summarize_records(
                location_fields_and_structure,
                f"L{gi + 1}",
                loc_identity,
                "Location fields + structure",
                "",
            )
        )

        for bi, gb in enumerate(gloc.get("buildings", []) or []):
            building_records = [
                r
                for r in all_records
                if (
                    r.location_idx == gi
                    and r.building_idx == bi
                )
            ]
            summary_rows.append(
                summarize_records(
                    building_records,
                    f"L{gi + 1} / B{bi + 1}",
                    building_identity(gb),
                    "Building",
                    (
                        "Includes building-parent membership check; "
                        "LocationNumber excluded."
                    ),
                )
            )

    for summary in summary_rows:
        for col, value in enumerate(summary, start=1):
            ws.cell(row=row, column=col, value=value)

        # Percentage formatting.
        ws.cell(row=row, column=8).number_format = "0.0000%"
        ws.cell(row=row, column=9).number_format = "0.0000%"

        # Overall row highlight.
        if summary[0] == "Overall":
            for col in range(1, 11):
                fill_cell(
                    ws.cell(row=row, column=col),
                    COLOR_CORRECT,
                )
        row += 1

    accuracy_summary_end = row - 1
    apply_table_borders(
        ws,
        accuracy_summary_start - 1,
        accuracy_summary_end,
        1,
        10,
    )
    row += 1

    # -------------------------------------------------------------------------
    # Values with issues
    # -------------------------------------------------------------------------
    ws.cell(row=row, column=1, value="Values with issues")
    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=9,
    )
    style_section_header(ws, row, 1, 9)
    row += 1

    issue_headers = [
        "Scope",
        "Location / building",
        "Check type",
        "Field / check",
        "Reference (ground truth)",
        "Value (ground truth)",
        "Value (prediction)",
        "Issue type",
        "Remarks",
    ]
    for col, value in enumerate(issue_headers, start=1):
        ws.cell(row=row, column=col, value=value)
    style_subheader(ws, row, 1, 9)
    row += 1

    issue_start = row

    issues = [
        r
        for r in all_records
        if r.status != "Correct"
    ]

    def issue_sort_key(record: Record) -> Tuple[int, int, int, str]:
        location = (
            -1
            if record.location_idx is None
            else record.location_idx
        )
        building = (
            -1
            if record.building_idx is None
            else record.building_idx
        )
        type_order = 0 if record.record_type == "Structure" else 1
        return (
            location,
            building,
            type_order,
            record.field,
        )

    issues.sort(key=issue_sort_key)

    for record in issues:
        values = [
            record.scope,
            record.identity,
            record.record_type,
            record.field,
            record.reference,
            display_value(
                record.field,
                record.ground_truth,
            ),
            display_value(
                record.field,
                record.prediction,
            ),
            record.status,
            issue_remark(record),
        ]

        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value)

        # IMPORTANT:
        # Do NOT color-highlight Issue type cells.
        row += 1

    issue_end = row - 1
    apply_table_borders(
        ws,
        issue_start - 1,
        issue_end,
        1,
        9,
    )

    # -------------------------------------------------------------------------
    # Accuracy sheet widths
    # -------------------------------------------------------------------------
    widths = {
        1: 18,
        2: 42,
        3: 18,
        4: 26,
        5: 42,
        6: 24,
        7: 34,
        8: 22,
        9: 42,
        10: 42,
    }
    set_column_widths(ws, widths)

    apply_global_styles(ws)

    # Restore section/subheader bold/fills after global formatting.
    # They remain at known row positions because we styled them while writing.
    # openpyxl preserves fill/font, so only generic alignment/font family needs
    # to be guaranteed.
    ws.freeze_panes = "A3"


# =============================================================================
# WORKBOOK GENERATION
# =============================================================================

def create_workbook(
    ground_truth_path: str | Path,
    prediction_path: str | Path,
    output_path: str | Path,
) -> Dict[str, Any]:
    gt = normalize_output_schema(
        load_json(ground_truth_path),
        fill_missing=True,
    )
    pred = normalize_output_schema(
        load_json(prediction_path),
        fill_missing=False,
    )

    (
        data_records,
        structural_records,
        gt_to_pred,
        pred_to_gt,
        building_matches,
        global_pred_buildings,
    ) = build_records(gt, pred)

    wb = Workbook()

    # Remove default sheet and create exactly two requested sheets.
    default_sheet = wb.active
    wb.remove(default_sheet)

    validation_ws = wb.create_sheet(SHEET_VALIDATION)
    accuracy_ws = wb.create_sheet(SHEET_ACCURACY)

    write_validation_sheet(
        validation_ws,
        gt,
        pred,
        gt_to_pred,
        building_matches,
    )

    write_accuracy_sheet(
        accuracy_ws,
        gt,
        pred,
        data_records,
        structural_records,
        gt_to_pred,
        pred_to_gt,
        building_matches,
        global_pred_buildings,
    )

    # Make Validation the first visible sheet.
    wb.active = 0

    output_path = Path(output_path)
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb.save(output_path)

    all_records = list(data_records) + list(structural_records)
    counts = Counter(r.status for r in all_records)
    expected = len(all_records)

    return {
        "output": str(output_path),
        "expected": expected,
        "correct": counts["Correct"],
        "formatting_issues": counts["Formatting Issue"],
        "wrong": counts["Wrong"],
        "accuracy_formatting_accepted": (
            (counts["Correct"] + counts["Formatting Issue"]) / expected
            if expected
            else 0.0
        ),
        "accuracy_strict": (
            counts["Correct"] / expected
            if expected
            else 0.0
        ),
        "locations_ground_truth": len(gt.get("propertyInfo", []) or []),
        "locations_prediction": len(pred.get("propertyInfo", []) or []),
    }


# =============================================================================
# CLI
# =============================================================================

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Compare prediction JSON against ground truth and generate "
            "Validation + Accuracy .xlsx workbook."
        )
    )
    parser.add_argument(
        "--ground-truth",
        required=True,
        help="Path to the ground-truth JSON file.",
    )
    parser.add_argument(
        "--prediction",
        required=True,
        help="Path to the prediction JSON file.",
    )
    parser.add_argument(
        "--output",
        required=True,
        help="Output .xlsx path.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    result = create_workbook(
        ground_truth_path=args.ground_truth,
        prediction_path=args.prediction,
        output_path=args.output,
    )

    # Minimal stdout is useful for execution logs.
    # The calling assistant should return only the .xlsx file to the user.
    print(result["output"])


if __name__ == "__main__":
    main()
